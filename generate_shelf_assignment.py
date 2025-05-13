import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import os

# File paths
SHELF_FILE = r"C:\Users\User\OneDrive - ensonmarket.com\shelf assignment\shelf information.xlsx"
FAMILY_FILE = r"C:\Users\User\OneDrive - ensonmarket.com\shelf assignment\family information.xlsx"
OUTPUT_FILE = r"C:\Users\User\OneDrive - ensonmarket.com\shelf assignment\Shelf_Assignment_Reversed_Output.xlsx"

def read_shelf_data(shelf_file):
    """Read shelf data from the input file and expand into individual shelves."""
    try:
        # Read the first sheet, specifying the columns we expect
        df = pd.read_excel(shelf_file, sheet_name=0)
        # Ensure expected columns are present
        expected_columns = ['section', 'aisles', 'sides', 'levels max', 'shelves max']
        missing_columns = [col for col in expected_columns if col not in df.columns]
        if missing_columns:
            print(f"Missing expected columns in shelf data: {missing_columns}")
            return None
        
        # Expand the summarized data into individual shelf entries
        expanded_data = []
        for _, row in df.iterrows():
            section = row['section']
            num_aisles = int(row['aisles'])
            num_sides = int(row['sides'])
            levels_max = int(row['levels max'])
            shelves_max = int(row['shelves max'])
            
            for aisle in range(1, num_aisles + 1):
                for side in range(1, num_sides + 1):
                    for level in range(1, levels_max + 1):
                        # Adjust shelves per level (assuming shelves_max is the max per level)
                        for shelf in range(1, shelves_max + 1):
                            expanded_data.append({
                                'Section': section,
                                'Aisle': aisle,
                                'Side': side,
                                'Level': level,
                                'Shelf': shelf
                            })
        
        expanded_df = pd.DataFrame(expanded_data)
        print(f"Read and expanded shelf data. Rows: {len(expanded_df)}")
        return expanded_df
    except Exception as e:
        print(f"Error reading shelf data: {str(e)}")
        return None

def read_family_data(family_file):
    """Read family data from the input file."""
    try:
        # Read all sheets to aggregate family and category data
        xls = pd.ExcelFile(family_file)
        families_dict = {}
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(family_file, sheet_name=sheet_name)
            # Assume family name is in cell A2 (row 2, column 1 in Excel, so index 0, 0 in pandas)
            family = str(df.iloc[0, 0]) if not pd.isna(df.iloc[0, 0]) else ""
            if family:
                # Categories are in row 2, starting from column B (index 1)
                categories = df.iloc[1, 1:].dropna().tolist()
                families_dict[family] = categories
        sub_categories = []
        for family, cats in families_dict.items():
            for cat in cats:
                sub_categories.append((family, cat))
        print(f"Read family data. Families: {len(families_dict)}")
        return sub_categories, families_dict
    except Exception as e:
        print(f"Error reading family data: {str(e)}")
        return None, None

def generate_output_file(shelf_data, sub_categories, families_dict, output_file):
    """Generate the output Excel file with dropdowns."""
    try:
        # Prepare the output DataFrame
        output_df = shelf_data.copy()
        output_df['Family'] = ""
        output_df['Category'] = ""
        
        # Write the initial output file
        output_df.to_excel(output_file, index=False)
        print(f"Initial output file created at: {output_file}")
        
        # Load the workbook with openpyxl to add dropdowns
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Get the list of families and all categories
        family_list = ",".join([f for f, _ in sub_categories])
        all_categories = set()
        for _, cats in families_dict.items():
            all_categories.update(cats)
        category_list = ",".join(all_categories) if all_categories else "No Categories Available"
        
        # Add data validation for each row
        last_row = ws.max_row
        for row in range(2, last_row + 1):  # Start from row 2 (after header)
            # Family dropdown in column F (6th column)
            dv_family = DataValidation(type="list", formula1=f'"{family_list}"', allow_blank=True)
            dv_family.add(f"F{row}")
            ws.add_data_validation(dv_family)
            
            # Category dropdown in column G (7th column)
            dv_category = DataValidation(type="list", formula1=f'"{category_list}"', allow_blank=True)
            dv_category.add(f"G{row}")
            ws.add_data_validation(dv_category)
        
        # Save the workbook with dropdowns
        wb.save(output_file)
        print(f"Dropdowns added to output file. Rows processed: {last_row - 1}")
    except Exception as e:
        print(f"Error generating output file: {str(e)}")
        raise

def save_updated_data(output_file):
    """Read the output file, preserve user selections, and save back."""
    try:
        # Read the current output file
        df = pd.read_excel(output_file)
        print(f"Read updated output file. Rows: {len(df)}")
        
        # Overwrite the file with the updated data (preserving dropdowns)
        df.to_excel(output_file, index=False)
        print(f"Updated data saved back to: {output_file}")
        
        # Re-add dropdowns using openpyxl
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Rebuild family and category lists (assuming original data is still available)
        _, families_dict = read_family_data(FAMILY_FILE)
        family_list = ",".join(families_dict.keys())
        all_categories = set()
        for cats in families_dict.values():
            all_categories.update(cats)
        category_list = ",".join(all_categories) if all_categories else "No Categories Available"
        
        # Add data validation for each row
        last_row = ws.max_row
        for row in range(2, last_row + 1):
            dv_family = DataValidation(type="list", formula1=f'"{family_list}"', allow_blank=True)
            dv_family.add(f"F{row}")
            ws.add_data_validation(dv_family)
            
            dv_category = DataValidation(type="list", formula1=f'"{category_list}"', allow_blank=True)
            dv_category.add(f"G{row}")
            ws.add_data_validation(dv_category)
        
        wb.save(output_file)
        print(f"Dropdowns re-added after saving updated data")
    except Exception as e:
        print(f"Error saving updated data: {str(e)}")
        raise

def main():
    """Main function to generate the output file and optionally save updated data."""
    # Validate input files
    if not os.path.exists(SHELF_FILE):
        print(f"Shelf file not found: {SHELF_FILE}")
        return
    if not os.path.exists(FAMILY_FILE):
        print(f"Family file not found: {FAMILY_FILE}")
        return
    
    # Read input data
    shelf_data = read_shelf_data(SHELF_FILE)
    if shelf_data is None:
        return
    
    sub_categories, families_dict = read_family_data(FAMILY_FILE)
    if sub_categories is None:
        return
    
    # Generate the output file with dropdowns
    generate_output_file(shelf_data, sub_categories, families_dict, OUTPUT_FILE)
    
    # Optionally save updated data (uncomment to use after making selections)
    # print("Make your selections in the output file, then press Enter to save changes.")
    # input()
    # save_updated_data(OUTPUT_FILE)

if __name__ == "__main__":
    main()